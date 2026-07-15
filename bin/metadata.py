#!/usr/bin/env python
"""
metadata.py — per-sample metadata capture for IRMA submission headers.

The point of this module is the user's core request: take the bare sequences
IRMA assembles and re-write their headers into a `<sample>-submission.fasta`
whose deflines carry the proper influenza strain name built from sample
metadata. Metadata can be supplied two interchangeable ways, both matched on the
sample name:

  1. Web entry — a small per-sample table saved as `sample_metadata.json`.
  2. Excel     — `sample_metadata.xlsx` (the same columns), so a lab can manage
                 it locally; the GUI offers Download / Replace for that file.

Both live in `<project>/irma/` so any tool/run sees the same metadata. The JSON
is the canonical store the web UI edits; the xlsx is generated from it on demand
and re-imported when a user uploads an edited copy. A run can also point at an
arbitrary external Excel (`--metadata-xlsx`), matched on a chosen key column.

Canonical fields (all optional except `sample`):
  sample, organism, host, state, collection_year, passage, subtype
The strain name is always auto-built as `A/host/state/sample/year(subtype)`;
any field left blank becomes `unknown_host`, `unknown_state`, `unknown_year`,
`unknown_subtype`, so headers are still well-formed when metadata is sparse.
`subtype` (e.g. H5N1) is prefilled from IRMA's HA/NA call but user-editable.

This module is pure-stdlib + openpyxl and is imported by both the pipeline
(`bin/`) and the FastAPI backend (which adds `bin/` to sys.path).
"""

from __future__ import annotations

import json
import re
from pathlib import Path
from typing import Any, Dict, List, Optional

# Canonical metadata fields, in display/column order.
FIELDS = ["sample", "organism", "host", "state", "collection_year", "passage", "subtype"]
DEFAULT_ORGANISM = "Influenza A virus"

METADATA_JSON = "sample_metadata.json"
METADATA_XLSX = "sample_metadata.xlsx"

# When importing an arbitrary user Excel, map its headers to our fields by
# matching any of these (case-insensitive, non-alphanumerics ignored).
_HEADER_ALIASES = {
    "sample": ["sample", "sample_id", "sampleid", "isolate", "name", "accession",
               "nvsl accession", "nvslaccession", "specimen"],
    "organism": ["organism", "virus", "agent"],
    "host": ["host", "species", "animal", "source"],
    "state": ["state", "location", "province", "region", "geo", "geographic location"],
    "collection_year": ["collection_year", "collectionyear", "year",
                         "collection year", "collection date", "collection_date", "date"],
    "passage": ["passage", "passage history", "passage_history"],
    "subtype": ["subtype", "sub type", "serotype", "hn", "h/n", "hn subtype",
                "hnsubtype", "strain"],
}

# Influenza-A segment metadata for descriptive submission deflines.
SEGMENT_NUMBER = {"PB2": 1, "PB1": 2, "PA": 3, "HA": 4, "NP": 5, "NA": 6, "MP": 7, "NS": 8}
# Canonical NCBI gene-product descriptions for the submission defline.
GENE_FULL_NAME = {
    "PB2": "polymerase PB2 (PB2)",
    "PB1": "polymerase PB1 (PB1)",
    "PA": "polymerase PA (PA)",
    "HA": "hemagglutinin (HA)",
    "NP": "nucleoprotein (NP)",
    "NA": "neuraminidase (NA)",
    "MP": "matrix protein 1 (M1) and matrix protein 2 (M2) (MP)",
    "NS": "nonstructural protein 1 (NS1) and nuclear export protein (NEP) (NS)",
}


def _norm(s: str) -> str:
    return re.sub(r"[^a-z0-9]+", "", str(s or "").strip().lower())


def _blank_record(sample: str) -> Dict[str, str]:
    rec = {f: "" for f in FIELDS}
    rec["sample"] = sample
    rec["organism"] = DEFAULT_ORGANISM
    return rec


# ---------------------------------------------------------------------------
# JSON store (canonical, edited by the web UI)
# ---------------------------------------------------------------------------
def load_json(irma_dir: Path) -> Dict[str, Dict[str, str]]:
    """Return {sample -> record} from sample_metadata.json (empty if absent)."""
    path = Path(irma_dir) / METADATA_JSON
    try:
        data = json.loads(path.read_text(encoding="utf-8"))
    except (OSError, ValueError):
        return {}
    out: Dict[str, Dict[str, str]] = {}
    if isinstance(data, dict):
        for sample, rec in data.items():
            if isinstance(rec, dict):
                norm = _blank_record(sample)
                for k, v in rec.items():
                    if k in FIELDS:
                        norm[k] = "" if v is None else str(v)
                norm["sample"] = sample
                out[sample] = norm
    return out


def save_json(irma_dir: Path, records: Dict[str, Dict[str, str]]) -> Path:
    irma_dir = Path(irma_dir)
    irma_dir.mkdir(parents=True, exist_ok=True)
    path = irma_dir / METADATA_JSON
    clean: Dict[str, Dict[str, str]] = {}
    for sample, rec in (records or {}).items():
        norm = _blank_record(sample)
        for k in FIELDS:
            if k in rec and rec[k] is not None:
                norm[k] = str(rec[k])
        norm["sample"] = sample
        clean[sample] = norm
    path.write_text(json.dumps(clean, indent=2, sort_keys=True) + "\n", encoding="utf-8")
    return path


# ---------------------------------------------------------------------------
# Excel I/O
# ---------------------------------------------------------------------------
def _resolve_columns(header: List[str]) -> Dict[str, int]:
    """Map our canonical fields -> column index in an arbitrary header row."""
    norm_hdr = [_norm(h) for h in header]
    mapping: Dict[str, int] = {}
    for field, aliases in _HEADER_ALIASES.items():
        for alias in aliases:
            na = _norm(alias)
            if na in norm_hdr:
                mapping[field] = norm_hdr.index(na)
                break
    return mapping


def read_xlsx(path: Path, key_col: Optional[str] = None) -> Dict[str, Dict[str, str]]:
    """Read a metadata workbook -> {sample -> record}, matched on the sample
    column (auto-detected, or `key_col` if given)."""
    from openpyxl import load_workbook

    wb = load_workbook(str(path), read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return {}
    header = ["" if c is None else str(c) for c in rows[0]]
    mapping = _resolve_columns(header)
    if key_col:
        nk = _norm(key_col)
        norm_hdr = [_norm(h) for h in header]
        if nk in norm_hdr:
            mapping["sample"] = norm_hdr.index(nk)
    if "sample" not in mapping:
        return {}
    out: Dict[str, Dict[str, str]] = {}
    for row in rows[1:]:
        cells = ["" if c is None else str(c).strip() for c in row]
        si = mapping["sample"]
        if si >= len(cells) or not cells[si]:
            continue
        sample = cells[si]
        rec = _blank_record(sample)
        for field, idx in mapping.items():
            if idx < len(cells) and cells[idx]:
                rec[field] = cells[idx]
        rec["sample"] = sample
        if not rec.get("organism"):
            rec["organism"] = DEFAULT_ORGANISM
        out[sample] = rec
    return out


def write_xlsx(records: Dict[str, Dict[str, str]], path: Path) -> Path:
    """Write {sample -> record} as a labeled metadata workbook."""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill

    wb = Workbook()
    ws = wb.active
    ws.title = "metadata"
    header_fill = PatternFill("solid", fgColor="4C8C8A")
    header_font = Font(bold=True, color="FFFFFF")
    for c, field in enumerate(FIELDS, start=1):
        cell = ws.cell(row=1, column=c, value=field)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="left")
    r = 2
    for sample in sorted(records):
        rec = records[sample]
        for c, field in enumerate(FIELDS, start=1):
            ws.cell(row=r, column=c, value=rec.get(field, ""))
        r += 1
    for c, field in enumerate(FIELDS, start=1):
        ws.column_dimensions[ws.cell(row=1, column=c).column_letter].width = max(12, len(field) + 4)
    ws.freeze_panes = "A2"
    Path(path).parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(path))
    return Path(path)


# ---------------------------------------------------------------------------
# Lookup (used by the pipeline)
# ---------------------------------------------------------------------------
def get_metadata(
    sample: str,
    irma_dir: Optional[Path] = None,
    xlsx: Optional[Path] = None,
    key_col: Optional[str] = None,
) -> Dict[str, str]:
    """Resolve one sample's metadata. Precedence: an explicit external Excel
    (`xlsx`) wins, else the project's sample_metadata.json. Always returns a
    full record (blanks where unknown) so header building never KeyErrors."""
    rec: Optional[Dict[str, str]] = None
    if xlsx:
        try:
            table = read_xlsx(Path(xlsx), key_col=key_col)
            rec = table.get(sample) or _match_loose(table, sample)
        except Exception:  # noqa: BLE001 — missing/locked file must not kill a run
            rec = None
    if rec is None and irma_dir:
        table = load_json(irma_dir)
        rec = table.get(sample) or _match_loose(table, sample)
    if rec is None:
        rec = _blank_record(sample)
    return rec


def _match_loose(table: Dict[str, Dict[str, str]], sample: str) -> Optional[Dict[str, str]]:
    """Fall back to a normalized-name match (e.g. trailing _S1 lane tags)."""
    ns = _norm(sample)
    for k, v in table.items():
        if _norm(k) == ns:
            return v
    for k, v in table.items():
        nk = _norm(k)
        if nk and (ns.startswith(nk) or nk.startswith(ns)):
            return v
    return None


# ---------------------------------------------------------------------------
# Strain name + submission deflines
# ---------------------------------------------------------------------------
def strain_string(rec: Dict[str, str], subtype: Optional[str] = None) -> str:
    """Build the influenza strain string from a record.

    Always assembles `A/host/state/sample/year(subtype)`. Blank fields become
    `unknown_host`, `unknown_state`, `unknown_year`, so the header is still
    well-formed when metadata is sparse. The subtype comes from the record's
    `subtype` field if set, else the `subtype` argument (IRMA's HA/NA call).
    """
    host = (rec.get("host") or "").strip() or "unknown_host"
    state = (rec.get("state") or "").strip() or "unknown_state"
    sample = (rec.get("sample") or "").strip() or "unknown_sample"
    year = (rec.get("collection_year") or "").strip() or "unknown_year"
    base = f"A/{host}/{state}/{sample}/{year}"
    # Subtype precedence: explicit metadata field, else IRMA's call.
    eff = (rec.get("subtype") or "").strip() or (subtype or "").strip()
    organism = (rec.get("organism") or DEFAULT_ORGANISM)
    is_flu = "influenza" in organism.lower()
    if eff and re.match(r"^H\d", eff) and "?" not in eff:
        # A clean influenza-style (HxNx) subtype.
        base = f"{base}({eff})"
    elif is_flu:
        # Influenza sample with no usable subtype call yet.
        base = f"{base}(unknown_subtype)"
    # else: a non-influenza module (e.g. CoV) — no (HxNx) suffix.
    return base


def submission_defline(rec: Dict[str, str], gene: str, subtype: Optional[str],
                       style: str = "ncbi") -> str:
    """Return a defline (without the leading '>') for one segment record.

    style="ncbi": GenBank-submission style with [organism=...] and a segment /
    gene-product description (the form labs paste into BankIt / GenBank).
    style="strain": a terse `strain_GENE` identifier.
    """
    gene = (gene or "").upper()
    strain = strain_string(rec, subtype)
    organism = (rec.get("organism") or DEFAULT_ORGANISM).strip()
    if style == "strain":
        return f"{strain}_{gene}"
    segnum = SEGMENT_NUMBER.get(gene)
    # NCBI-submission style: `SeqN [organism=...](strain) segment N, <product> gene, complete cds.`
    seq_id = f"Seq{segnum}" if segnum else f"{strain}_{gene}"
    seg_txt = f"segment {segnum}, " if segnum else ""
    # Influenza segments get a "<product> gene"; other genomes just name the gene.
    desc = f"{GENE_FULL_NAME[gene]} gene" if gene in GENE_FULL_NAME else gene
    return f"{seq_id} [organism={organism}]({strain}) {seg_txt}{desc}, complete cds."
